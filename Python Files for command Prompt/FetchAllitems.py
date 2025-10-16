import requests
import logging
import os
from openpyxl import Workbook, load_workbook

# === Credentials ===
CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
REFRESH_TOKEN = "1000.89243e274b1607f4b1b23760417ca19c.ba478a62cdc32ce464c074ec84b6b38a"
ORG_ID = "891730368"

logging.basicConfig(level=logging.INFO, format="%(message)s")

# === Get Access Token ===
def get_access_token():
    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": REFRESH_TOKEN,
        "client_id": CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    r = requests.post(url, params=params)
    r.raise_for_status()
    token = r.json()["access_token"]
    logging.info(f"[INFO] Access token received: {token[:15]}...")
    return token

# === Save results to Excel ===
def save_to_excel(data, filename="Fetche All item_log.xlsx"):
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Item ID", "SKU", "Name"])

    for row in data:
        ws.append(row)

    wb.save(filename)
    logging.info(f"[INFO] Data saved to {filename}")

# === Fetch all inventory items ===
def fetch_inventory_items():
    token = get_access_token()
    page = 1
    per_page = 200
    all_items = []

    while True:
        url = "https://inventory.zoho.com/api/v1/items"
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        params = {
            "organization_id": ORG_ID,
            "page": page,
            "per_page": per_page
        }
        r = requests.get(url, headers=headers, params=params)
        r.raise_for_status()
        items = r.json().get("items", [])

        if not items:
            break

        for item in items:
            item_id = item.get("item_id", "")
            sku = item.get("sku", "")
            name = item.get("name", "")

            # Log to console
            logging.info(f"ID: {item_id} | SKU: {sku} | Name: {name}")

            all_items.append([item_id, sku, name])

        page += 1

    save_to_excel(all_items)

# === Run ===
if __name__ == "__main__":
    fetch_inventory_items()
