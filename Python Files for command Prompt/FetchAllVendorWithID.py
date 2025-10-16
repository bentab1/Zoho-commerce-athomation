import requests
from datetime import datetime, timezone
from openpyxl import Workbook, load_workbook
import os

# === Zoho OAuth details ===
ZOHO_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_REFRESH_TOKEN = "1000.89243e274b1607f4b1b23760417ca19c.ba478a62cdc32ce464c074ec84b6b38a"
ORG_ID = 891730368

EXCEL_FILENAME = "FetchItemwithVendordetailsInlcuding ID&Numbe.xlsx"
SHEET_NAME_DATA = "Inventory"
SHEET_NAME_LOGS = "Logs"

# === Logging ===
def log(message, log_ws):
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    log_ws.append([timestamp, message])
    print(f"{timestamp} | {message}")

def get_or_create_workbook(filename):
    return load_workbook(filename) if os.path.exists(filename) else Workbook()

def get_or_create_worksheet(wb, sheet_name):
    return wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

# === Zoho OAuth token refresh ===
def refresh_access_token():
    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": ZOHO_REFRESH_TOKEN,
        "client_id": ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, params=params)
    resp.raise_for_status()
    return resp.json()["access_token"]

# === Fetch full item details ===
def fetch_item_details(access_token, item_id):
    url = f"https://www.zohoapis.com/inventory/v1/items/{item_id}?organization_id={ORG_ID}"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    resp = requests.get(url, headers=headers)
    resp.raise_for_status()
    return resp.json().get("item", {})

# === Fetch vendor details ===
def fetch_vendor_details(access_token, vendor_id):
    if not vendor_id:
        return {}
    url = f"https://www.zohoapis.com/inventory/v1/contacts/{vendor_id}?organization_id={ORG_ID}"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    resp = requests.get(url, headers=headers)
    if resp.status_code == 200:
        return resp.json().get("contact", {})
    return {}

# === Fetch all inventory items ===
def fetch_inventory_items(access_token):
    items = []
    page = 1
    per_page = 200
    while True:
        url = f"https://www.zohoapis.com/inventory/v1/items?organization_id={ORG_ID}&page={page}&per_page={per_page}"
        headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        batch = resp.json().get("items", [])
        if not batch:
            break
        items.extend(batch)
        page += 1
    return items

# === Main ===
def main():
    wb = get_or_create_workbook(EXCEL_FILENAME)
    log_ws = get_or_create_worksheet(wb, SHEET_NAME_LOGS)
    data_ws = get_or_create_worksheet(wb, SHEET_NAME_DATA)

    log("▶️ Script started.", log_ws)

    try:
        access_token = refresh_access_token()
    except Exception as e:
        log(f"❌ Token refresh failed: {e}", log_ws)
        wb.save(EXCEL_FILENAME)
        return

    headers = [
        "item_id", "item_name", "sku", "description", "rate", "purchase_rate",
        "available_stock", "reorder_level", "brand",
        "vendor_id", "vendor_number",
        "vendor_display_name", "vendor_company_name", "vendor_email", "vendor_phone"
    ]
    if data_ws.max_row == 0:
        data_ws.append(headers)
    else:
        if [cell.value for cell in data_ws[1]] != headers:
            data_ws.delete_rows(1)
            data_ws.append(headers)
        if data_ws.max_row > 1:
            data_ws.delete_rows(2, data_ws.max_row)

    vendor_cache = {}

    try:
        items = fetch_inventory_items(access_token)
        log(f"📦 Total inventory items fetched: {len(items)}", log_ws)

        for i, item in enumerate(items, start=1):
            vendor_display = vendor_company = vendor_email = vendor_phone = ""
            vendor_id = vendor_number = ""
            brand = item.get("brand", "")

            # Fetch full item details to get vendor_id
            try:
                item_details = fetch_item_details(access_token, item.get("item_id"))
                vendor_id = item_details.get("preferred_vendor_id") or item_details.get("vendor_id", "")

                if vendor_id:
                    if vendor_id in vendor_cache:
                        cached = vendor_cache[vendor_id]
                        vendor_display = cached["vendor_display"]
                        vendor_company = cached["vendor_company"]
                        vendor_email = cached["vendor_email"]
                        vendor_phone = cached["vendor_phone"]
                        vendor_number = cached["vendor_number"]
                    else:
                        contact = fetch_vendor_details(access_token, vendor_id)
                        vendor_display = contact.get("contact_name", "")
                        vendor_company = contact.get("company_name", "")
                        vendor_email = contact.get("email", "")
                        vendor_phone = contact.get("phone", "")
                        vendor_number = contact.get("vendor_number", "")  # Confirm correct key here
                        vendor_cache[vendor_id] = {
                            "vendor_display": vendor_display,
                            "vendor_company": vendor_company,
                            "vendor_email": vendor_email,
                            "vendor_phone": vendor_phone,
                            "vendor_number": vendor_number
                        }
                else:
                    log(f"⚠️ No vendor for item: {item.get('name')}", log_ws)
            except Exception as e:
                log(f"❌ Error fetching details for item_id {item.get('item_id')}: {e}", log_ws)

            row = [
                item.get("item_id", ""),
                item.get("name", ""),
                item.get("sku", ""),
                item.get("description", ""),
                item.get("rate", ""),
                item.get("purchase_rate", ""),
                item.get("available_stock", ""),
                item.get("reorder_level", ""),
                brand,
                vendor_id,
                vendor_number,
                vendor_display,
                vendor_company,
                vendor_email,
                vendor_phone
            ]
            data_ws.append(row)

            if i % 10 == 0:
                log(f"Progress: {i}/{len(items)} items processed", log_ws)

        log("✅ Finished fetching inventory with vendor info.", log_ws)

    except Exception as e:
        log(f"❌ Error fetching inventory: {e}", log_ws)

    wb.save(EXCEL_FILENAME)
    log("▶️ Script ended.", log_ws)

if __name__ == "__main__":
    main()
