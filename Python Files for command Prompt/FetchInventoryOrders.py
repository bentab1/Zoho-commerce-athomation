import requests
from datetime import datetime
from openpyxl import Workbook, load_workbook
import os

# Zoho OAuth details
ZOHO_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_REFRESH_TOKEN = "1000.89243e274b1607f4b1b23760417ca19c.ba478a62cdc32ce464c074ec84b6b38a"
ORG_ID = 891730368  # Replace with your actual organization ID

EXCEL_FILENAME = "fetchAll_inventoryOrders.ipynb.xlsx"
SHEET_NAME_DATA = "InventoryOrders"

def log(message):
    timestamp = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    print(f"{timestamp} | {message}")

def get_or_create_workbook(filename):
    if os.path.exists(filename):
        wb = load_workbook(filename)
    else:
        wb = Workbook()
    return wb

def get_or_create_worksheet(wb, sheet_name):
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
    return ws

def refresh_access_token():
    # Implement your token refresh logic here
    # Example (you must replace with your own refresh logic):
    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": ZOHO_REFRESH_TOKEN,
        "client_id": ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type": "refresh_token",
    }
    response = requests.post(url, params=params)
    response.raise_for_status()
    access_token = response.json().get("access_token")
    if not access_token:
        raise Exception("Failed to refresh access token")
    return access_token

def fetch_sales_orders(access_token):
    # Fetch sales orders list from Zoho Inventory API
    url = f"https://inventory.zoho.com/api/v1/salesorders?organization_id={ORG_ID}&page=1&per_page=200"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    return data.get("salesorders", [])

def fetch_order_details(access_token, salesorder_id):
    url = f"https://inventory.zoho.com/api/v1/salesorders/{salesorder_id}?organization_id={ORG_ID}"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    return data.get("salesorder", {})

def fetch_item_details(access_token, item_id):
    url = f"https://inventory.zoho.com/api/v1/items/{item_id}?organization_id={ORG_ID}"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    return data.get("item", {})

def fetch_vendor_details(access_token, vendor_id):
    url = f"https://inventory.zoho.com/api/v1/contacts/{vendor_id}?organization_id={ORG_ID}"
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    response = requests.get(url, headers=headers)
    response.raise_for_status()
    data = response.json()
    return data.get("contact", {})

def main():
    wb = get_or_create_workbook(EXCEL_FILENAME)
    data_ws = get_or_create_worksheet(wb, SHEET_NAME_DATA)

    log("▶️ Script started.")

    try:
        access_token = refresh_access_token()
    except Exception as e:
        log(f"❌ Script stopped due to failure in token refresh: {e}")
        wb.save(EXCEL_FILENAME)
        return

    # Clear all existing rows including headers before writing new data
    data_ws.delete_rows(1, data_ws.max_row)

    # Append headers fresh
    headers = [
        "salesorder_id", "salesorder_number", "date", "customer_name",
        "status", "total", "created_time", "last_modified_time",
        "item_id", "item_name", "sku",
        "vendor_display_name", "vendor_company_name", "vendor_email", "vendor_phone", "brand"
    ]
    data_ws.append(headers)

    vendor_cache = {}

    try:
        orders = fetch_sales_orders(access_token)
        log(f"📦 Total Sales Orders fetched: {len(orders)}")

        for order in orders:
            salesorder_id = order.get("salesorder_id")
            try:
                order_details = fetch_order_details(access_token, salesorder_id)
                line_items = order_details.get("line_items", [])

                for item in line_items:
                    item_id = item.get("item_id", "")
                    item_name = item.get("name", "")
                    sku = item.get("sku", "")

                    vendor_display = ""
                    vendor_company = ""
                    vendor_email = ""
                    vendor_phone = ""
                    brand = ""

                    if item_id:
                        if item_id in vendor_cache:
                            cached = vendor_cache[item_id]
                            vendor_display = cached["vendor_display"]
                            vendor_company = cached["vendor_company"]
                            vendor_email = cached["vendor_email"]
                            vendor_phone = cached["vendor_phone"]
                            brand = cached["brand"]
                        else:
                            try:
                                item_data = fetch_item_details(access_token, item_id)
                                brand = item_data.get("brand", "")
                                vendor_id = item_data.get("vendor_id")

                                if vendor_id:
                                    contact = fetch_vendor_details(access_token, vendor_id)
                                    vendor_display = contact.get("contact_name", "")
                                    vendor_company = contact.get("company_name", "")
                                    vendor_email = contact.get("email", "")
                                    vendor_phone = contact.get("phone", "")
                                else:
                                    log(f"⚠️ No vendor_id for item: {item_name}")

                                vendor_cache[item_id] = {
                                    "vendor_display": vendor_display,
                                    "vendor_company": vendor_company,
                                    "vendor_email": vendor_email,
                                    "vendor_phone": vendor_phone,
                                    "brand": brand
                                }
                            except Exception as e:
                                log(f"❌ Error fetching item/vendor details for item_id {item_id}: {e}")

                    row = [
                        order_details.get("salesorder_id", ""),
                        order_details.get("salesorder_number", ""),
                        order_details.get("date", ""),
                        order_details.get("customer_name", ""),
                        order_details.get("status", ""),
                        order_details.get("total", ""),
                        order_details.get("created_time", ""),
                        order_details.get("last_modified_time", ""),
                        item_id,
                        item_name,
                        sku,
                        vendor_display,
                        vendor_company,
                        vendor_email,
                        vendor_phone,
                        brand
                    ]

                    data_ws.append(row)

            except Exception as e:
                log(f"❌ Error fetching details for sales order {salesorder_id}: {e}")

        log("✅ Finished fetching sales orders with vendor info.")

    except Exception as e:
        log(f"❌ Error fetching sales order list: {e}")

    log("▶️ Script ended.")
    wb.save(EXCEL_FILENAME)


if __name__ == "__main__":
    main()
