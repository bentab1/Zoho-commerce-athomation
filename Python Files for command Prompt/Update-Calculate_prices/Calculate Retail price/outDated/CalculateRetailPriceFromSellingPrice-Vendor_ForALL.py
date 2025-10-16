import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import time

# -------------------------------
# CONFIGURATION
# -------------------------------
ZOHO_COM_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_COM_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_COM_REFRESH_TOKEN = "1000.e6b8fed22697a4847fa2912eb7bcdc71.db25a2b68ef38e7bf01858ad17489870"

ZOHO_INV_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_INV_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_INV_REFRESH_TOKEN = "1000.89243e274b1607f4b1b23760417ca19c.ba478a62cdc32ce464c074ec84b6b38a"

ORG_ID = "891730368"
NGN_RATE = 1533.98908652912
ADJUSTMENT = 1
LOG_FILE = "calculateUpdate_retailPrice_fromSellingPrice_log.xlsx"

# -------------------------------
# TOKENS
# -------------------------------
def get_zoho_commerce_token():
    url = "https://accounts.zoho.com/oauth/v2/token"
    payload = {
        "refresh_token": ZOHO_COM_REFRESH_TOKEN,
        "client_id": ZOHO_COM_CLIENT_ID,
        "client_secret": ZOHO_COM_CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, data=payload).json()
    if "access_token" in resp:
        return resp["access_token"]
    raise Exception(f"❌ Failed to get Commerce token: {resp}")

def get_zoho_inventory_token():
    url = "https://accounts.zoho.com/oauth/v2/token"
    params = {
        "refresh_token": ZOHO_INV_REFRESH_TOKEN,
        "client_id": ZOHO_INV_CLIENT_ID,
        "client_secret": ZOHO_INV_CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, params=params)
    if resp.status_code == 200:
        print("✅ Inventory access token obtained.")
        return resp.json().get("access_token")
    else:
        raise Exception(f"❌ Failed to get Inventory token: {resp.text}")

# -------------------------------
# INVENTORY FUNCTIONS
# -------------------------------
def fetch_all_vendors(token):
    vendors = []
    page = 1
    while True:
        url = f"https://inventory.zoho.com/api/v1/contacts?page={page}&per_page=200&type=vendor&organization_id={ORG_ID}"
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        resp = requests.get(url, headers=headers).json()
        if "contacts" not in resp:
            break
        vendors.extend([v for v in resp["contacts"] if v.get("contact_type") == "vendor"])
        if not resp.get("page_context", {}).get("has_more_page", False):
            break
        page += 1
    return vendors

def select_vendor(token):
    vendors = fetch_all_vendors(token)
    print("\nAvailable Vendors:")
    for i, v in enumerate(vendors, 1):
        print(f"{i}. {v['contact_name']} (ID: {v['contact_id']})")
    choice = input("Select vendor (number or ID): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(vendors):
        return vendors[int(choice)-1]['contact_id']
    for v in vendors:
        if v["contact_id"] == choice:
            return choice
    return None

def fetch_inventory_items(token, vendor_id=None):
    items = []
    page = 1
    while True:
        url = f"https://inventory.zoho.com/api/v1/items?organization_id={ORG_ID}&page={page}&per_page=200"
        if vendor_id:
            url += f"&vendor_id={vendor_id}"
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        resp = requests.get(url, headers=headers).json()
        if "items" not in resp:
            break
        for i in resp["items"]:
            items.append({"sku": i.get("sku","").upper(), "name": i.get("name","")})
        if not resp.get("page_context", {}).get("has_more_page", False):
            break
        page += 1
    return items

# -------------------------------
# COMMERCE FUNCTIONS
# -------------------------------
def wait_and_retry(response):
    if response.status_code == 429:
        retry_after = int(response.headers.get("Retry-After", 5))
        print(f"⚠️ Rate limit reached. Retrying after {retry_after} seconds...")
        time.sleep(retry_after)
        return True
    return False

def fetch_zoho_commerce_products(access_token):
    page = 1
    products = []
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}
    while True:
        url = f"https://commerce.zoho.com/store/api/v1/products?page={page}&per_page=200&organization_id={ORG_ID}"
        response = requests.get(url, headers=headers)
        while wait_and_retry(response):
            response = requests.get(url, headers=headers)
        if response.status_code != 200:
            break
        data = response.json()
        if not data.get("products"):
            break
        products.extend(data["products"])
        if data.get("page_context", {}).get("has_more_page"):
            page += 1
        else:
            break
    return products

def calculate_new_retail_price(current_selling):
    try:
        current_selling = float(current_selling or 0.0)
    except:
        current_selling = 0.0
    return round(current_selling * 1.2121)

def log_to_excel(log_entries):
    if os.path.exists(LOG_FILE):
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Datetime","Product Name","Variant Name","SKU","Selling Price","New Retail","NGN Rate","Adjustment","Status"]
        ws.append(headers)
        for i, col in enumerate(headers,1):
            ws.column_dimensions[get_column_letter(i)].width=22
            ws.cell(row=1,column=i).font=Font(bold=True)
            ws.cell(row=1,column=i).alignment=Alignment(horizontal='center')
    for entry in log_entries:
        ws.append(entry)
    wb.save(LOG_FILE)

# -------------------------------
# UPDATE FUNCTION
# -------------------------------
def update_commerce_prices():
    inv_token = get_zoho_inventory_token()
    com_token = get_zoho_commerce_token()

    mode = input("Choose mode:\n1 - Calculate by Vendor\n2 - Calculate for All\nEnter 1 or 2: ").strip()
    if mode == "1":
        vendor_id = select_vendor(inv_token)
        inventory_items = fetch_inventory_items(inv_token, vendor_id)
    else:
        inventory_items = fetch_inventory_items(inv_token)

    commerce_products = fetch_zoho_commerce_products(com_token)

    # Build Commerce SKU map
    commerce_sku_map = {}
    for p in commerce_products:
        if p.get("variants"):
            for v in p["variants"]:
                commerce_sku_map[v.get("sku","").upper()] = (p["product_id"], v)
        else:
            commerce_sku_map[p.get("sku","").upper()] = (p["product_id"], p)

    headers = {
        "Authorization": f"Zoho-oauthtoken {com_token}",
        "Content-Type": "application/json",
        "X-com-zoho-commerce-organizationid": ORG_ID
    }

    log_entries = []

    for item in inventory_items:
        sku = item.get("sku","").upper()
        if sku not in commerce_sku_map:
            continue  # skip if SKU not in Commerce
        product_id, com_variant = commerce_sku_map[sku]
        selling_price = com_variant.get("rate") or 0
        if selling_price==0:
            continue
        new_retail = calculate_new_retail_price(selling_price)

        payload = {"variants":[{"variant_id": com_variant.get("variant_id"),"label_rate": new_retail}]}
        update_url = f"https://commerce.zoho.com/store/api/v1/products/{product_id}?organization_id={ORG_ID}"
        resp = requests.put(update_url, headers=headers, json=payload)
        while wait_and_retry(resp):
            resp = requests.put(update_url, headers=headers, json=payload)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        status = "Success" if resp.status_code in range(200,300) else f"Failed ({resp.status_code})"
        print(f"{status} - SKU: {sku}, Selling: {selling_price}, New Retail: {new_retail}")
        log_entries.append([timestamp, item.get("name",""), com_variant.get("name",""), sku, selling_price, new_retail, NGN_RATE, ADJUSTMENT, status])

    log_to_excel(log_entries)
    print(f"\n📦 Total variants processed: {len(log_entries)}")
# -------------------------------
# MAIN
# -------------------------------
if __name__ == "__main__":
    while True:
        update_commerce_prices()

        again = input("\n🔄 Do you want to perform another update? (y/n): ").strip().lower()
        if again != "y":
            print("👋 Exiting program. Goodbye!")
            break

