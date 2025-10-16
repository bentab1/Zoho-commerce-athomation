import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import time

# -------------------------------
# CONFIGURATION
# -------------------------------
ZOHO_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_REFRESH_TOKEN = "1000.e6b8fed22697a4847fa2912eb7bcdc71.db25a2b68ef38e7bf01858ad17489870"
ORG_ID = "891730368"
NGN_RATE = 1533.98908652912
ADJUSTMENT = 1
LOG_FILE = "calculateUpdate_retailPrice_fromSellingPriceSKURequest_log.xlsx"

# -------------------------------
# FUNCTIONS
# -------------------------------

def get_sku_list():
    mode = input("Choose SKU input mode:\n1 - Generate SKUs\n2 - Provide SKU list (comma separated)\nEnter choice (1 or 2): ").strip()
    if mode == "1":
        prefix = input("Enter prefix (e.g. AKASKU1): ").strip()
        item_name = input("Enter item name (e.g. IP): ").strip()
        start = int(input("Enter start number (e.g. 1): ").strip())
        end = int(input("Enter end number (e.g. 10): ").strip())
        mid_code = item_name.upper()[:2]

        pad_option = input("Choose padding type:\n1 - Fixed 6-digit zero padding (e.g. 000065)\n2 - No padding (e.g. 563)\n3 - Dynamic padding based on input\nEnter choice (1, 2, or 3): ").strip()

        if pad_option == "1":
            pad_len = 6
            skus = [f"{prefix}-{mid_code}-{str(num).zfill(pad_len)}" for num in range(start, end + 1)]
        elif pad_option == "2":
            skus = [f"{prefix}-{mid_code}-{str(num)}" for num in range(start, end + 1)]
        elif pad_option == "3":
            pad_len = max(len(str(start)), len(str(end)))
            skus = [f"{prefix}-{mid_code}-{str(num).zfill(pad_len)}" for num in range(start, end + 1)]
        else:
            print("Invalid padding option. Exiting.")
            exit()

        print(f"Generated SKUs: {skus}")
        return skus

    elif mode == "2":
        sku_raw = input("Enter SKUs separated by commas: ").strip()
        skus = [sku.strip().upper() for sku in sku_raw.split(",") if sku.strip()]
        print(f"Using provided SKUs: {skus}")
        return skus
    else:
        print("Invalid mode. Exiting.")
        exit()

def get_zoho_access_token():
    """Get Zoho OAuth access token using refresh token."""
    url = "https://accounts.zoho.com/oauth/v2/token"
    payload = {
        "refresh_token": ZOHO_REFRESH_TOKEN,
        "client_id": ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, data=payload)
    data = resp.json()
    if "access_token" in data:
        return data["access_token"]
    else:
        raise Exception(f"❌ Failed to get access token: {data}")

def calculate_new_retail_price(current_selling):
    """Calculate retail price as selling price × 1.2121."""
    try:
        current_selling = float(current_selling or 0.0)
    except:
        current_selling = 0.0

    new_retail = current_selling * 1.2121
    return round(new_retail)

def wait_and_retry(response):
    """Handle Zoho API rate limits by waiting and retrying."""
    if response.status_code == 429:
        retry_after = int(response.headers.get("Retry-After", 5))
        print(f"⚠️ Rate limit reached. Retrying after {retry_after} seconds...")
        time.sleep(retry_after)
        return True
    return False

def fetch_zoho_commerce_products(access_token):
    """Fetch all products from Zoho Commerce."""
    page = 1
    products = []
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}

    while True:
        url = f"https://commerce.zoho.com/store/api/v1/products?page={page}&per_page=200&organization_id={ORG_ID}"
        response = requests.get(url, headers=headers)

        while wait_and_retry(response):
            response = requests.get(url, headers=headers)

        if response.status_code != 200:
            print(f"❌ Failed to fetch products (Page {page}): {response.status_code} - {response.text}")
            break

        data = response.json()
        if not data.get("products"):
            break

        products.extend(data["products"])

        if data.get("page_context", {}).get("has_more_page"):
            page += 1
        else:
            break

    return products

def log_to_excel(log_entries):
    """Log price updates to Excel."""
    if os.path.exists(LOG_FILE):
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Datetime", "Product Name", "Variant Name", "SKU", "Selling Price",
            "New Retail (Retail Price)", "NGN Rate", "Adjustment", "Status"
        ]
        ws.append(headers)
        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
            ws.cell(row=1, column=i).font = Font(bold=True)
            ws.cell(row=1, column=i).alignment = Alignment(horizontal='center')

    for entry in log_entries:
        ws.append(entry)

    wb.save(LOG_FILE)

def update_commerce_retail_prices_for_skus(ngn_rate, adjustment, sku_list):
    """Update retail prices only for variants matching SKUs in sku_list."""

    access_token = get_zoho_access_token()
    products = fetch_zoho_commerce_products(access_token)

    headers = {
        "Authorization": f"Zoho-oauthtoken {access_token}",
        "Content-Type": "application/json",
        "X-com-zoho-commerce-organizationid": ORG_ID
    }
    log_entries = []

    # To track which SKUs were matched and updated
    matched_skus = set()

    for product in products:
        product_name = product.get("document_name", "Unknown Product")
        product_id = product.get("product_id")
        variants = product.get("variants", [])

        for variant in variants:
            variant_name = variant.get("name", "Unknown Variant")
            sku = variant.get("sku", "").strip()
            selling_price = variant.get("rate", 0)
            variant_id = variant.get("variant_id")

            if not sku:
                continue

            # Check if this variant's SKU is in our SKU list (case-insensitive match)
            if sku.upper() not in map(str.upper, sku_list):
                continue

            matched_skus.add(sku.upper())

            if selling_price is None or selling_price == 0:
                print(f"⚠️ Skipping {variant_name} ({sku}) because selling_price is zero or missing.")
                status = "Skipped (No selling price)"
                timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                log_entries.append([
                    timestamp, product_name, variant_name, sku, selling_price,
                    "", ngn_rate, adjustment, status
                ])
                continue

            new_retail = calculate_new_retail_price(selling_price)

            payload = {
                "variants": [
                    {
                        "variant_id": variant_id,
                        "label_rate": new_retail
                    }
                ]
            }

            update_url = f"https://commerce.zoho.com/store/api/v1/products/{product_id}?organization_id={ORG_ID}"
            resp = requests.put(update_url, headers=headers, json=payload)

            while wait_and_retry(resp):
                resp = requests.put(update_url, headers=headers, json=payload)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            if resp.status_code in range(200, 300):
                print(f"✅ Updated {variant_name} ({sku}): Selling Price ₦{selling_price} → Retail Price ₦{new_retail}")
                status = "Success"
            else:
                print(f"❌ Failed to update {variant_name} ({sku}): {resp.status_code} - {resp.text}")
                status = f"Failed ({resp.status_code})"

            log_entries.append([
                timestamp, product_name, variant_name, sku, selling_price,
                new_retail, ngn_rate, adjustment, status
            ])

    # Log SKUs not found
    not_found_skus = [sku for sku in sku_list if sku.upper() not in matched_skus]
    if not_found_skus:
        print("\n⚠️ The following SKUs were NOT found in Zoho Commerce and were NOT updated:")
        for nf_sku in not_found_skus:
            print(f" - {nf_sku}")

        # Add to log with "Not Found" status
        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for nf_sku in not_found_skus:
            log_entries.append([
                timestamp, "", "", nf_sku, "", "", ngn_rate, adjustment, "Not Found"
            ])

    log_to_excel(log_entries)
    print(f"\n📦 Total variants processed: {len(log_entries)}")
    return log_entries

# -------------------------------
# MAIN
# -------------------------------
if __name__ == "__main__":
    sku_list = get_sku_list()  # use your enhanced SKU input function

    update_commerce_retail_prices_for_skus(NGN_RATE, ADJUSTMENT, sku_list)
