import requests
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import time

# -------------------------------
# CONFIGURATION
# -------------------------------
ZOHO_CLIENT_ID = "1000.S0MCEJ7YY6TVO6HC3RED00QLK7KDRA"
ZOHO_CLIENT_SECRET = "1bbb7b39b57c784a38cbdc9c6e82496ee690bde86d"
ZOHO_REFRESH_TOKEN = "1000.e6b8fed22697a4847fa2912eb7bcdc71.db25a2b68ef38e7bf01858ad17489870"
ZOHO_INV_REFRESH_TOKEN = "1000.89243e274b1607f4b1b23760417ca19c.ba478a62cdc32ce464c074ec84b6b38a"
ORG_ID = "891730368"
BASE_RATE = 1533.98908652912
NGN_RATE = 1533.98908652912
## To increase price with adjustment, divide what to increase by the current price to get the increament. 
## E.g. If we have 2000 as the current price and we want to add 500 to our price .: 500/2000 = 0.25 increament
## Therefore, our adjustment = 0.25. If I want to remove 500 from the current price that, adjustment = -0.25
##The script uses exchange rate to adjsut the price if the adjsutment is not equals 0
ADJUSTMENT = 0
LOG_FILE = "updateCommerce_sellingPrice_retailPriceWithSKURequest_log.xlsx"

# -------------------------------
# UTILITY FUNCTIONS
# -------------------------------

def get_zoho_access_token(refresh_token=ZOHO_REFRESH_TOKEN):
    url = "https://accounts.zoho.com/oauth/v2/token"
    payload = {
        "refresh_token": refresh_token,
        "client_id": ZOHO_CLIENT_ID,
        "client_secret": ZOHO_CLIENT_SECRET,
        "grant_type": "refresh_token"
    }
    resp = requests.post(url, data=payload)
    data = resp.json()
    if "access_token" in data:
        return data["access_token"]
    else:
        raise Exception(f"❌ Failed to get access token: {data}")

def wait_and_retry(response):
    if response.status_code == 429:
        retry_after = int(response.headers.get("Retry-After", 5))
        print(f"⚠️ Rate limit reached. Retrying after {retry_after} seconds...")
        time.sleep(retry_after)
        return True
    return False


def calculate_new_prices(current_selling, current_retail, ngn_rate, base_rate, adjustment):
    try:
        current_selling = float(current_selling or 0.0)
    except:
        current_selling = 0.0

    try:
        current_retail = float(current_retail or 0.0)
    except:
        current_retail = 0.0

    if adjustment != 0:
        # Apply adjustment as a percent increase
        new_selling = current_selling * (1 + adjustment)
        new_retail = current_retail * (1 + adjustment)
        percentagePrice_diff = adjustment * 100  # convert to %
    else:
        # Exchange rate based calculation
        percentagePrice_diff = ((ngn_rate / base_rate) - 1) * 100  # convert to %
        new_selling = current_selling * (1 + percentagePrice_diff / 100)
        new_retail = current_retail * (1 + percentagePrice_diff / 100)

    print(f"Price change: {percentagePrice_diff:.2f}%")
    return round(new_selling), round(new_retail)



def fetch_zoho_commerce_products(access_token):
    page = 1
    products = []
    headers = {"Authorization": f"Zoho-oauthtoken {access_token}"}

    while True:
        url = f"https://commerce.zoho.com/store/api/v1/products?page={page}&per_page=200&organization_id={ORG_ID}"
        response = requests.get(url, headers=headers)
        while wait_and_retry(response):
            response = requests.get(url, headers=headers)
        if response.status_code != 200:
            print(f"❌ Failed to fetch products (Page {page}): {response.status_code} - {response.text}")
            break

        data = response.json()
        if not data.get("products"):
            break

        products.extend(data["products"])
        if data.get("page_context", {}).get("has_more_page"):
            page += 1
        else:
            break
    return products

# -------------------------------
# INVENTORY VENDOR FUNCTIONS
# -------------------------------

def fetch_all_vendors(token):
    vendors = []
    page = 1
    while True:
        url = f"https://inventory.zoho.com/api/v1/contacts?page={page}&per_page=200&type=vendor&organization_id={ORG_ID}"
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        resp = requests.get(url, headers=headers).json()
        if "contacts" not in resp:
            break
        vendors.extend([v for v in resp["contacts"] if v.get("contact_type") == "vendor"])
        if not resp.get("page_context", {}).get("has_more_page", False):
            break
        page += 1
    return vendors

def select_vendor(token):
    vendors = fetch_all_vendors(token)
    print("\nAvailable Vendors:")
    for i, v in enumerate(vendors, 1):
        print(f"{i}. {v['contact_name']} (ID: {v['contact_id']})")
    choice = input("Select vendor (number or ID): ").strip()
    if choice.isdigit() and 1 <= int(choice) <= len(vendors):
        return vendors[int(choice)-1]['contact_id']
    for v in vendors:
        if v["contact_id"] == choice:
            return choice
    return None

def fetch_inventory_items_by_vendor(vendor_id, token):
    items = []
    page = 1
    while True:
        url = f"https://inventory.zoho.com/api/v1/items?organization_id={ORG_ID}&vendor_id={vendor_id}&per_page=200&page={page}"
        headers = {"Authorization": f"Zoho-oauthtoken {token}"}
        resp = requests.get(url, headers=headers)
        resp.raise_for_status()
        data = resp.json()
        items_page = data.get("items", [])
        if not items_page:
            break
        items.extend([i.get("sku", "").upper() for i in items_page if i.get("sku")])
        if not data.get("page_context", {}).get("has_more_page"):
            break
        page += 1
    return items

# -------------------------------
# SKU INPUT FUNCTIONS
# -------------------------------

def get_sku_list_manual_or_generated():
    mode = input("Choose SKU input mode:\n1 - Generate SKUs\n2 - Provide SKU list (comma separated)\nEnter choice (1 or 2): ").strip()
    if mode == "1":
        prefix = input("Enter prefix (e.g. AKASKU1): ").strip()
        item_name = input("Enter item name (e.g. IP): ").strip()
        start = int(input("Enter start number (e.g. 1): ").strip())
        end = int(input("Enter end number (e.g. 10): ").strip())
        mid_code = item_name.upper()[:2]

        pad_option = input("Choose padding type:\n1 - Fixed 6-digit zero padding\n2 - No padding\n3 - Dynamic padding\nEnter choice: ").strip()
        if pad_option == "1":
            skus = [f"{prefix}-{mid_code}-{str(num).zfill(6)}" for num in range(start, end+1)]
        elif pad_option == "2":
            skus = [f"{prefix}-{mid_code}-{num}" for num in range(start, end+1)]
        elif pad_option == "3":
            pad_len = max(len(str(start)), len(str(end)))
            skus = [f"{prefix}-{mid_code}-{str(num).zfill(pad_len)}" for num in range(start, end+1)]
        else:
            print("Invalid padding option. Exiting.")
            exit()
        return skus
    elif mode == "2":
        sku_raw = input("Enter SKUs separated by commas: ").strip()
        return [sku.strip().upper() for sku in sku_raw.split(",") if sku.strip()]
    else:
        print("Invalid mode. Exiting.")
        exit()

# -------------------------------
# EXCEL LOGGING
# -------------------------------

def log_to_excel(log_entries):
    if os.path.exists(LOG_FILE):
        wb = openpyxl.load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Datetime", "Product Name", "SKU", "Old Selling", "Old Retail",
            "New Selling", "New Retail", "NGN Rate", "Adjustment", "Status"
        ]
        ws.append(headers)
        for i, col in enumerate(headers, 1):
            ws.column_dimensions[get_column_letter(i)].width = 22
            ws.cell(row=1, column=i).font = Font(bold=True)
            ws.cell(row=1, column=i).alignment = Alignment(horizontal='center')

    for entry in log_entries:
        ws.append(entry)

    wb.save(LOG_FILE)

# -------------------------------
# MAIN UPDATE FUNCTION
# -------------------------------

def update_commerce_prices(ngn_rate, adjustment, skus=None, vendor_mode=False):
    access_token = get_zoho_access_token()
    products = fetch_zoho_commerce_products(access_token)
    headers = {
        "Authorization": f"Zoho-oauthtoken {access_token}",
        "Content-Type": "application/json",
        "X-com-zoho-commerce-organizationid": ORG_ID
    }

    log_entries = []
    matched_skus = set()

    for product in products:
        product_id = product.get("product_id")
        name = product.get("name", "")
        variants = product.get("variants", [])

        # Single variant products
        if not variants:
            sku = product.get("sku", "").upper()
            if skus is not None and sku not in skus:
                continue
            matched_skus.add(sku)

            current_selling = product.get("rate", 0)
            current_retail = product.get("label_rate", current_selling)
            new_selling, new_retail = calculate_new_prices(current_selling, current_retail, ngn_rate, BASE_RATE, adjustment)

            payload = {"rate": new_selling, "label_rate": new_retail}
            update_url = f"https://commerce.zoho.com/store/api/v1/products/{product_id}?organization_id={ORG_ID}"
            resp = requests.put(update_url, headers=headers, json=payload)
            while wait_and_retry(resp):
                resp = requests.put(update_url, headers=headers, json=payload)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            status = "Success" if resp.status_code in range(200, 300) else f"Failed ({resp.status_code})"
            log_entries.append([timestamp, name, sku, current_selling, current_retail, new_selling, new_retail, ngn_rate, adjustment, status])
            print(f"{'✅' if status=='Success' else '❌'} Updated {name} ({sku}): Selling ₦{current_selling} → ₦{new_selling}, Retail ₦{current_retail} → ₦{new_retail}")
            continue

        # Multiple variants
        for variant in variants:
            sku = variant.get("sku", "").upper()
            if skus is not None and sku not in skus:
                continue
            matched_skus.add(sku)

            variant_name = variant.get("name", "Unknown Variant")
            current_selling = variant.get("rate") or 0
            current_retail = variant.get("label_rate") or current_selling
            if current_selling == 0:
                print(f"⚠️ Skipping variant {variant_name} ({sku}) due to missing selling price")
                continue

            new_selling, new_retail = calculate_new_prices(current_selling, current_retail, ngn_rate, BASE_RATE, adjustment)
            payload = {"variants": [{"variant_id": variant.get("variant_id"), "rate": new_selling, "label_rate": new_retail}]}

            update_url = f"https://commerce.zoho.com/store/api/v1/products/{product_id}?organization_id={ORG_ID}"
            resp = requests.put(update_url, headers=headers, json=payload)
            while wait_and_retry(resp):
                resp = requests.put(update_url, headers=headers, json=payload)

            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            status = "Success" if resp.status_code in range(200, 300) else f"Failed ({resp.status_code})"
            log_entries.append([timestamp, f"{name} - {variant_name}", sku, current_selling, current_retail, new_selling, new_retail, ngn_rate, adjustment, status])
            print(f"{'✅' if status=='Success' else '❌'} Updated variant {variant_name} ({sku}): Selling ₦{current_selling} → ₦{new_selling}, Retail ₦{current_retail} → ₦{new_retail}")

    # Log SKUs not found
    if skus is not None:
        not_found = [sku for sku in skus if sku not in matched_skus]
        for nf in not_found:
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            log_entries.append([timestamp, "", nf, "", "", "", "", ngn_rate, adjustment, "Not Found"])
            print(f"⚠️ SKU Not Found: {nf}")

    log_to_excel(log_entries)
    print(f"\n📦 Total SKUs processed: {len(log_entries)}")
    return log_entries

# -------------------------------
# MAIN SCRIPT
# -------------------------------
if __name__ == "__main__":
    while True:
        mode = input("Choose mode:\n1 - By SKU list\n2 - By Vendor\n3 - All Products\nEnter choice: ").strip()
        
        if mode == "1":
            sku_list = get_sku_list_manual_or_generated()
            update_commerce_prices(NGN_RATE, ADJUSTMENT, skus=sku_list)
        elif mode == "2":
            inv_token = get_zoho_access_token(ZOHO_INV_REFRESH_TOKEN)
            vendor_id = select_vendor(inv_token)
            if not vendor_id:
                print("❌ No vendor selected. Exiting this run.")
                continue
            sku_list = fetch_inventory_items_by_vendor(vendor_id, inv_token)
            print(f"⚡ Fetched {len(sku_list)} SKUs for Vendor {vendor_id}")
            update_commerce_prices(NGN_RATE, ADJUSTMENT, skus=sku_list)
        elif mode == "3":
            print("⚡ Processing ALL products (no SKU/Vendor filter)")
            update_commerce_prices(NGN_RATE, ADJUSTMENT, skus=None)
        else:
            print("Invalid mode. Exiting this run.")
            continue
        
        again = input("\n🔄 Do you want to perform another price update? (y/n): ").strip().lower()
        if again not in ("y", "yes"):
            print("👋 Exiting. No further updates will be performed.")
            break
